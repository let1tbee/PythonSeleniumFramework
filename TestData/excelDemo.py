import openpyxl

file_path = "C:\\Users\\ratus\\PycharmProjects\\pythonTest\\PythonAQA\\PythonSelFramework\\TestData\\excelDemo.xlsx"
book = openpyxl.load_workbook(file_path ) #excel file location
sheet = book.active    # picks an active sheet
cell = sheet.cell(row=1,column=2) #picks cell
print(cell.value) #grabs value
sheet.cell(row=2,column=2).value = "Oleksii" #write data to excel
print(sheet.cell(row=2,column=2).value)

print(sheet.max_row) #prints number of rows which are has maximum values
print(sheet.max_column) # same with columns

print(sheet['A2'].value) #other way to call a cell

bool.save(file_path) #saves book
dic = {}
for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value == "Testcase2":
        for j in range(2,sheet.max_column+1):
            dic[sheet.cell(row=1,column=j).value] = sheet.cell(row=i,column=j).value

