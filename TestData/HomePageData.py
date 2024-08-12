import openpyxl




class HomePageData:
    test_HomePage_data = {}

    book = openpyxl.load_workbook(
        "C:\\Users\\ratus\\PycharmProjects\\pythonTest\\PythonAQA\\PythonSelFramework\\TestData\\excelDemo.xlsx")  # excel file location
    sheet = book.active  # picks an active sheet
    cell = sheet.cell(row=1, column=2)  # picks cell

    @staticmethod #declare as a static method to be able use it without creating class object
    def getTestData(testcase_name):    #as static method declared no more need in self
        for i in range(1, HomePageData.sheet.max_row + 1):
            if HomePageData.sheet.cell(row=i, column=1).value == testcase_name:
                for j in range(2, HomePageData.sheet.max_column + 1):
                    HomePageData.test_HomePage_data[HomePageData.sheet.cell(row=1, column=j).value] = HomePageData.sheet.cell(row=i, column=j).value

        print( [HomePageData.test_HomePage_data]) #should be returned as list but not as a dictionary
        return [HomePageData.test_HomePage_data]